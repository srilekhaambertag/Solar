import pandas as pd
import numpy as np
import os
import glob
import Config as cn
import ftplib
import datetime, re, os
import openpyxl
import pdb
class SaveFiles:
    def __init__(self):
        self.save()
        pass
    
    def save(self):
        try:
            os.makedirs('Storage')
        except:
            pass
        
        try:
            os.makedirs('Results')
        except:
            pass
        
        
        try:
            os.makedirs(os.path.join(cn.StoragePath,'LocalWMS'))
        except:
            pass
        
        try:
            os.makedirs(os.path.join(cn.StoragePath,'MeterOne'))
        except:
            pass
        
        try:
            os.makedirs(os.path.join(cn.StoragePath,'MeterTwo'))
        except:
            pass
        
        try:
            os.makedirs(os.path.join(cn.StoragePath,'ExternalWMS'))
        except:
            pass
        
        try:
            os.makedirs(os.path.join(cn.StoragePath,'PreProcessedData'))
        except:
            pass
        #file = open(cn.filesave,'w')
        #file.close()
        
    def saveLocalWMSFromFTP(self, downloadall = False):
        ftp = ftplib.FTP(cn.ftpAddress) 
        ftp.login(cn.ftplogin,cn.ftppasswd) 
        ftp.cwd('/MBScada')

        if downloadall:
            print('Local WMS folder empty, downloading all files: \n')
            fileslist = ftp.nlst()   
            localWmslist = [ filelist for filelist in fileslist if 'Barod_WMSData' in filelist]
            for file in localWmslist:
                path = os.path.join(cn.LocalWMSPath,file)
                ftp.retrbinary("RETR " + file, open(path, 'wb').write)
                print("File is available")
        else:
            #Time=str(datetime.datetime.now() - datetime.timedelta(minutes=15))[11:16]
            
            Current_Date = str(pd.Timestamp(datetime.date.today()))[0:10]
            Current_Date = re.sub("-", "_", Current_Date)
            file = 'Todays Clean Energy,Barod_WMSData_'+Current_Date+'.xls'
            path = os.path.join(cn.LocalWMSPath,file)
            try:
                os.remove(path)
            except:
                pass
            print('removed local WMS file for today {}\n'.format(Current_Date))
            print('Downloading updated Local WMS file\n')
            ftp.retrbinary("RETR " + file, open(path, 'wb').write)
            print('Downloaded local WMS file for today {}\n'.format(Current_Date))
        pass
    
    
    def savePowerFromFTP(self, downloadall = False, power = 'BarodCSVMeter_', savePath = cn.MeterOnePath,
                         filename = 'Todays Clean Energy,BarodCSVMeter_Data_'):
        ftp = ftplib.FTP(cn.ftpAddress) 
        ftp.login(cn.ftplogin,cn.ftppasswd) 
        ftp.cwd('/MBScada')

        if downloadall:
            print('Power folder empty, downloading all files: \n')
            fileslist = ftp.nlst()   
            Powerlist = [ filelist for filelist in fileslist if power in filelist]
            for file in Powerlist:
                path = os.path.join(savePath,file)
                ftp.retrbinary("RETR " + file, open(path, 'wb').write)
                print("File is available")
        else:
            #Time=str(datetime.datetime.now() - datetime.timedelta(minutes=15))[11:16]
            Current_Date = str(pd.Timestamp(datetime.date.today()))[0:10]
            Current_Date = re.sub("-", "_", Current_Date)
            file = filename+Current_Date+'.csv'
            path = os.path.join(savePath,file)
            try:
                os.remove(path)
            except:
                pass
            print('removed power file for today {}\n'.format(Current_Date))
            print('Downloading updated Power file\n')
            ftp.retrbinary("RETR " + file, open(path, 'wb').write)
            print('Downloaded power file for today {}\n'.format(Current_Date))
           
        
        pass
    
    
    
    def saveForecastedPower(self, mergedDf, ypred, loc = '/Barod_Location1_'):
        self.datetime = str(datetime.datetime.now())[:16]
        filenameLoc = cn.saveResultsPath +loc+ self.datetime +'.xlsx'
        filenameLoc = filenameLoc.replace(" ","_")
        filenameLoc = filenameLoc.replace("-","_")
        filenameLoc = filenameLoc.replace(":","_")

        if loc =='/Barod_Location1_':
            xfilelocation = openpyxl.load_workbook(cn.location1TemplateFilepath)
        else:
            xfilelocation = openpyxl.load_workbook(cn.location2TemplateFilepath)
        
        sheet = xfilelocation.get_sheet_by_name('Sheet1')
        revision = {'05:00':'1','06:30':'2','08:00':'3','09:30':'4','11:00':'5',
                            '12:30':'6','14:00':'7','15:30':'8','17:00':'9','18:30':'10'}
        time = self.datetime[11:16]
        sheet['B3'] = str(datetime.datetime.now())[:10]
        try:
            revisionNum = revision.get(time)
            sheet['B4'] = "R"+revisionNum
        except:
            sheet['B4'] = "R"
            
            
        sheet['B5'] = str(datetime.datetime.now())[11:16]
        
        time1 = datetime.time(0,0)
        time2 = mergedDf.loc[0,'DateTime'].time()
        addInd = int((datetime.datetime.combine(datetime.date.today(), time2) - datetime.datetime.combine(datetime.date.today(), time1))/datetime.timedelta(0,900))
        startingRow = 10+addInd
        #pdb.set_trace()
        j = 0
        starttime = datetime.time(6,00)
        endtime = datetime.time(18,00)
        #pdb.set_trace()
        for i in range(startingRow,startingRow+len(ypred)+3):
            #print(i)
            #pdb.set_trace()
            t = datetime.datetime.strptime(sheet['B'+str(i)].value, '%H:%M').time()
            
            if (t <= starttime) or (t >= endtime):
                sheet['E'+str(i)] = str(0)
            else:
                sheet['E'+str(i)] = str(round(ypred[j],2))
            j+=1
            
        '''for i in range(10,105):
            if datetime.datetime.strptime(sheet['B'+str(i)].value, '%H:%M').time() < time2:
                sheet['E'+str(i)] = str(0)
        
        for i in range(startingRow+len(ypred),startingRow+len(ypred)+3):
            sheet['E'+str(i)] = str(0)'''
            
        
        xfilelocation.save(filenameLoc)
        pass
        
    
    def saveForecastedPowerFromExternalInps(self, mergedDf, loc = '/Barod_Location1_'):
        ypred = mergedDf['Forecasted'].values
        self.datetime = str(datetime.datetime.now())[:16]
        filenameLoc = cn.saveResultsPathForExternal +loc+ self.datetime +'.xlsx'
        filenameLoc = filenameLoc.replace(" ","_")
        filenameLoc = filenameLoc.replace("-","_")
        filenameLoc = filenameLoc.replace(":","_")

        if loc =='/Barod_Location1_':
            xfilelocation = openpyxl.load_workbook(cn.location1TemplateFilepath)
        else:
            xfilelocation = openpyxl.load_workbook(cn.location2TemplateFilepath)
        
        sheet = xfilelocation.get_sheet_by_name('Sheet1')
        revision = {'05:00':'1','06:30':'2','08:00':'3','09:30':'4','11:00':'5',
                            '12:30':'6','14:00':'7','15:30':'8','17:00':'9','18:30':'10'}
        time = self.datetime[11:16]
        sheet['B3'] = str(datetime.datetime.now())[:10]
        try:
            revisionNum = revision.get(time)
            sheet['B4'] = "R"+revisionNum
        except:
            sheet['B4'] = "R"
            
            
        sheet['B5'] = str(datetime.datetime.now())[11:16]
        
        time1 = datetime.time(0,0)
        time2 = mergedDf.loc[0,'DateTime'].time()
        addInd = int((datetime.datetime.combine(datetime.date.today(), time2) - datetime.datetime.combine(datetime.date.today(), time1))/datetime.timedelta(0,900))
        startingRow = 10+addInd
        #pdb.set_trace()
        j = 0
        starttime = datetime.time(6,00)
        endtime = datetime.time(18,00)
        #pdb.set_trace()
        for i in range(startingRow,startingRow+len(ypred)+3):
            #print(i)
            #pdb.set_trace()
            t = datetime.datetime.strptime(sheet['B'+str(i)].value, '%H:%M').time()
            
            if (t <= starttime) or (t >= endtime):
                sheet['E'+str(i)] = str(0)
            else:
                sheet['E'+str(i)] = str(round(ypred[j],2))
            j+=1
            
        '''for i in range(10,105):
            if datetime.datetime.strptime(sheet['B'+str(i)].value, '%H:%M').time() < time2:
                sheet['E'+str(i)] = str(0)
        
        for i in range(startingRow+len(ypred),startingRow+len(ypred)+3):
            sheet['E'+str(i)] = str(0)'''
            
        
        xfilelocation.save(filenameLoc)
        pass